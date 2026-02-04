import os
import pandas as pd
from datetime import datetime
from fpdf import FPDF
import storage
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def generate_exports(section_name):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename_base = f"{section_name.replace(' ', '_')}_{timestamp}"

    xlsx_path = os.path.join('exports', f"{filename_base}.xlsx")
    pdf_path = os.path.join('exports', f"{filename_base}.pdf")

    xlsx_res = None
    pdf_res = None

    try:
        generate_xlsx(section_name, xlsx_path)
        xlsx_res = xlsx_path
    except Exception as e:
        print(f"Error generating XLSX for {section_name}: {e}")

    try:
        generate_pdf(section_name, pdf_path)
        pdf_res = pdf_path
    except Exception as e:
        print(f"Error generating PDF for {section_name}: {e}")

    return xlsx_res, pdf_res

def generate_xlsx(section_name, xlsx_path):
    royal_blue_fill = PatternFill(start_color="002366", end_color="002366", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        for theme_id in range(1, 5):
            response = storage.get_response(section_name, theme_id)
            sheet_name = ["Budget", "Bureau", "Formation", "Salaries"][theme_id-1]

            if response:
                data = response['data']
                template = storage.get_template(theme_id, version=response['template_version'])

                rows = []
                row_types = [] # 0: normal, 1: header/title

                if template['structure']['type'] == 'budget':
                    for group in template['structure']['groups']:
                        row_types.append(1)
                        rows.append([group['title'], ""])
                        for field in group['fields']:
                            row_types.append(0)
                            rows.append([field['label'], data.get(field['id'], '')])
                        row_types.append(0)
                        rows.append(["", ""])

                elif template['structure']['type'] == 'fixed_table':
                    row_types.append(1)
                    header = ["Poste"] + template['structure']['cols']
                    rows.append(header)
                    for r_def in template['structure']['rows']:
                        row_types.append(0)
                        row = [r_def['label']]
                        for col in template['structure']['cols']:
                            key = f"{r_def['id']}_{col.lower()}"
                            row.append(data.get(key, ''))
                        rows.append(row)

                elif template['structure']['type'] == 'dynamic_table':
                    row_types.append(1)
                    header = [col['label'] for col in template['structure']['cols']]
                    rows.append(header)
                    for r_data in data.get('rows', []):
                        row_types.append(0)
                        row = [r_data.get(col['id'], '') for col in template['structure']['cols']]
                        rows.append(row)

                df = pd.DataFrame(rows)
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                ws = writer.sheets[sheet_name]

                # Column widths
                for i, col in enumerate(df.columns):
                    column_letter = get_column_letter(i + 1)
                    ws.column_dimensions[column_letter].width = 35

                for r_idx, (row, r_type) in enumerate(zip(ws.iter_rows(), row_types)):
                    if r_type == 1:
                        for cell in row:
                            cell.fill = royal_blue_fill
                            cell.font = white_font
                            cell.border = thin_border
                    else:
                        if any(c.value for c in row):
                            for cell in row:
                                cell.border = thin_border
            else:
                pd.DataFrame([["Aucune donnée saisie"]]).to_excel(writer, sheet_name=sheet_name, index=False, header=False)

def generate_pdf(section_name, pdf_path):
    def clean_txt(t):
        return str(t).replace("€", "EUR")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    for theme_id in range(1, 5):
        pdf.add_page()
        pdf.set_font("helvetica", 'B', 16)
        pdf.set_text_color(0, 35, 102) # Royal Blue

        titles = ["Budget prévisionnel", "Bureau directeur", "Diplômes et plan de formation", "Salariés"]
        title = titles[theme_id-1]

        pdf.cell(0, 10, f"{section_name} - {title}", align='C', new_x="LMARGIN", new_y="NEXT")
        pdf.ln(10)

        pdf.set_font("helvetica", '', 12)
        pdf.set_text_color(0, 0, 0)

        response = storage.get_response(section_name, theme_id)
        if response:
            data = response['data']
            template = storage.get_template(theme_id, version=response['template_version'])

            if template['structure']['type'] == 'budget':
                for group in template['structure']['groups']:
                    pdf.set_font("helvetica", 'B', 13)
                    pdf.cell(0, 10, group['title'], new_x="LMARGIN", new_y="NEXT")
                    pdf.set_font("helvetica", '', 11)
                    for field in group['fields']:
                        val = clean_txt(data.get(field['id'], ''))
                        # Robust multi_cell with explicit width and X reset
                        pdf.set_x(pdf.l_margin)
                        pdf.multi_cell(pdf.epw, 8, f"{clean_txt(field['label'])} : {val}")
                    pdf.ln(5)

            elif template['structure']['type'] == 'fixed_table':
                for r_def in template['structure']['rows']:
                    pdf.set_font("helvetica", 'B', 11)
                    pdf.set_x(pdf.l_margin)
                    pdf.multi_cell(pdf.epw, 8, clean_txt(r_def['label']))
                    pdf.set_font("helvetica", '', 11)
                    for col in template['structure']['cols']:
                        key = f"{r_def['id']}_{col.lower()}"
                        val = clean_txt(data.get(key, ''))
                        pdf.set_x(pdf.l_margin + 10)
                        pdf.multi_cell(pdf.epw - 10, 8, f"{clean_txt(col)}: {val}")
                    pdf.ln(2)

            elif template['structure']['type'] == 'dynamic_table':
                for i, r_data in enumerate(data.get('rows', [])):
                    pdf.set_font("helvetica", 'B', 11)
                    pdf.set_x(pdf.l_margin)
                    pdf.cell(0, 8, f"Ligne {i+1}", new_x="LMARGIN", new_y="NEXT")
                    pdf.set_font("helvetica", '', 11)
                    for col in template['structure']['cols']:
                        val = clean_txt(r_data.get(col['id'], ''))
                        pdf.set_x(pdf.l_margin + 10)
                        pdf.multi_cell(pdf.epw - 10, 8, f"{clean_txt(col['label'])}: {val}")
                    pdf.ln(2)
        else:
            pdf.cell(0, 10, "Aucune donnée saisie pour ce formulaire.", new_x="LMARGIN", new_y="NEXT")

    pdf.output(pdf_path)
